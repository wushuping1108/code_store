import requests
import unittest
import os
import logging
import time
import pymysql
import yagmail
from openpyxl import load_workbook
from selenium import  webdriver


# 使用类的方式来管理http请求的发送，它是爷爷类
class BaseHttp(object):

    # 统一http发送方式
    def sendHttp(self, path, method='post', **kwargs):
        url = '{}{}'.format(self.host, path)
        result = requests.request(method=method, url=url, **kwargs)
        return result

# 统一封装校验类
class VerifyClass(unittest.TestCase):
    # 校验状态码
    # 校验json格式响应体
    # text/html格式响应体
    # 校验响应体的特殊字段

    # 通过调用一个方法，去校验接口的多样性
    def verify_json_data(self, target, key, result_data):
        '''
        :param target:  未处理的接口响应
        :param key:  需要获取响应的Key
        :param result_data:  校验的字段
        :return:
        '''
        code = target.status_code
        target = target.json()
        self.assertEqual(200, code)
        self.assertEqual(target.get(key), result_data)

    def verify_html_data(self, target, result_data):
        code = target.status_code
        target = target.text
        self.assertEqual(200, code)
        self.assertIn(result_data, target)


# 当前文件路径
cur_path = os.path.dirname(os.path.realpath(__file__))

# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 类名称：InsertLog
# 类的目的：写日志
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：你爸爸
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------

# log_path是存放日志的路径
log_path = os.path.join(os.path.dirname(cur_path), 'log')
# 如果不存在这个logs文件夹，就自动创建一个
if not os.path.exists(log_path):
    os.mkdir(log_path)


class InsertLog():
    def __init__(self):
        # 文件的命名
        self.logname = os.path.join(log_path, '%s.log' % time.strftime('%Y_%m_%d'))
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.DEBUG)
        # 日志输出格式
        self.formatter = logging.Formatter(
            '[%(asctime)s - %(funcName)s line: %(lineno)3d] - %(levelname)s: %(message)s')

    def __console(self, level, message):
        # 创建一个FileHandler，用于写到本地
        fh = logging.FileHandler(self.logname, 'a')  # 追加模式  这个是python2的
        # fh = logging.FileHandler(self.logname, 'a', encoding='utf-8')  # 这个是python3的
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(self.formatter)
        self.logger.addHandler(fh)

        # 创建一个StreamHandler,用于输出到控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        ch.setFormatter(self.formatter)
        self.logger.addHandler(ch)

        if level == 'info':
            self.logger.info(message)
        elif level == 'debug':
            self.logger.debug(message)
        elif level == 'warning':
            self.logger.warning(message)
        elif level == 'error':
            self.logger.error(message)
        # 这两行代码是为了避免日志输出重复问题
        self.logger.removeHandler(ch)
        self.logger.removeHandler(fh)
        # 关闭打开的文件
        fh.close()

    def debug(self, message):
        self.__console('debug', message)

    def info(self, message):
        self.__console('info', message)

    def warning(self, message):
        self.__console('warning', message)

    def error(self, message):
        self.__console('error', message)


# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 函数/过程名称：ReadMySQLData
# 函数/过程的目的：写MySQL数据库单条数据
# 假设：无
# 影响：无
# 输入：无
# 返回值：查询一行数据（元组类型）
# 创建者：你爸爸
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------


# def ReadMySQLData(host, port, user, passwrod, db, sql):
#     try:
#         conn = MySQLdb.connect(host=host, \
#                                port=port, \
#                                user=user, \
#                                passwd=passwrod, \
#                                db=db)
#         curs = conn.cursor()
#         curs.execute(sql)
#         r = curs.fetchone()
#         # r = curs.fetchall()
#         curs.close()
#         conn.close()
#         return r
#     except BaseException as msg:
#         log = InsertLog()
#         log.error(msg)
#
#         # r = ReadMySQLData('192.168.1.123',3306,'edu','edu','edu',"select * from xsmart_users where username='15393701589'")
#         # print r


# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 函数/过程名称：GetNewReport
# 函数/过程的目的：获取最新报告文件
# 假设：无
# 影响：无
# 输入：无
# 返回值：文件全路径
# 创建者：你爸爸
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
FD = "./reports"


def GetNewReport(FileDir=FD):
    # 打印目录所在所有文件名（列表对象）
    l = os.listdir(FileDir)
    # 按时间排序
    l.sort(key=lambda fn: os.path.getmtime(FileDir + "\\" + fn))
    # 获取最新的文件保存到file_new
    f = os.path.join(FileDir, l[-1])
    return f


# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 类名称：SendEmail
# 类的目的：发送文本邮件或发送带附件邮件
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：你爸爸
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
def SendEmail(s_user, s_pwd, host, port, to_user, body, subject, report_file):
    """
    :param s_user: 发送者账号
    :param s_pwd:  发送者授权码
    :param host:   邮件服务器
    :param port:   非SSL端口25 SSL端口465 or 994
    :param to_user:  接收者账号，多个用列表表示
    :param body:  邮件接收内容
    :param subject:  邮件接收主题
    :param report_file:  测试报告
    :return:
    """
    send = yagmail.SMTP(user=s_user, password=s_pwd, host=host, port=port)
    # body = '老师您好，这是我今天晚上的昨天麻烦您清查，谢谢。'
    if type(to_user) is list:
        # 群发邮件，如果还是发送失败提示554，尝试把双方都加为发送联系人，cc为抄送自己
        send.send(to=to_user, cc=to_user, subject=subject, contents=[body, report_file])
        flag = True

    elif type(to_user) is str:
        # 发送给单个用户
        send.send(to=to_user, subject=subject, contents=[body, report_file])
        flag = True
    else:
        flag = False
    return flag

# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 类名称：create__browser_driver
# 类的目的：获取浏览器驱动
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：你爸爸
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------

#配置浏览器类型
bs = 'gc'

def create__browser_driver(b=bs):
    try:
        if b == 'gc':
            dv = webdriver.Chrome()
        elif b == 'ff':
            dv = webdriver.Firefox()
        elif b == 'ie':
            dv = webdriver.Ie()
        else:
            pass
        return dv
    except BaseException:
        pass


# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 类名称：getTestData
# 类的目的：获取测试数据
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：你爸爸
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------

# 获取测试数据
def getTestData(local, requirement, casename):
    wb = load_workbook(local)
    ws = wb[requirement]
    for i in range(len(ws['A'])):
        if ws['A'][i].value == casename:
            result = ws.cell(row=ws['A'][i].row,column=5).value
            if result is not None:
                stuff = result.find(',')
                if stuff != -1:
                    data = result.split(',')
                    return data
    wb.close()

# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 类名称：readMySQLData
# 类的目的：连接数据库
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：你爸爸
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
def readMySQLData(host, port, user, passwrod, db, sql):
    try:
        conn = pymysql.connect(host=host, \
                               port=port, \
                               user=user, \
                               passwd=passwrod, \
                               db=db)
        curs = conn.cursor()
        curs.execute(sql)
        r = curs.fetchone()
        curs.close()
        conn.close()
        return r
    except BaseException as msg:
        log = InsertLog()
        log.error(msg)

if __name__ == '__main__':
    ...